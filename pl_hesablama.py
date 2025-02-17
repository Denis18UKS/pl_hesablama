import sys
import os
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QGridLayout, QLineEdit, QLabel,
    QTableWidget, QTableWidgetItem, QPushButton, QDateEdit, QComboBox,
    QWidget, QFileDialog, QHBoxLayout, QSpinBox, QScrollArea, QFrame, QGraphicsPixmapItem, QGraphicsView, QGraphicsScene
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QPixmap
import openpyxl
import datetime
import subprocess
from PyQt5.QtGui import QIcon
from openpyxl.styles import Alignment, Font, Border, Side
from PyQt5.QtGui import QDoubleValidator
from PyQt5.QtWidgets import QStyledItemDelegate, QLineEdit


# Глобальные константы
IMAGES_FOLDER = "images/"

class NumericDelegate(QStyledItemDelegate):
    """Делегат для ограничения ввода только числами."""
    def __init__(self, parent=None):
        super().__init__(parent)

    def createEditor(self, parent, option, index):
        # Создаем редактор с валидатором
        editor = QLineEdit(parent)
        validator = QDoubleValidator(0.0, 1000000.0, 2, editor)  # Диапазон от 0 до 1000000, точность до 2 знаков
        validator.setNotation(QDoubleValidator.StandardNotation)
        editor.setValidator(validator)
        return editor

class OrderApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Форма заказа")
        self.setGeometry(100, 100, 1200, 800)

        # Создаем папку для хранения изображений
        if not os.path.exists(IMAGES_FOLDER):
            os.makedirs(IMAGES_FOLDER)

        # Главное виджет и макет
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout()
        self.central_widget.setLayout(self.main_layout)

        # Поля ввода информации
        self.create_order_info_section()

        # Таблица продуктов
        self.create_product_table()

        # Блок изображений
        self.create_image_section()

        # Блок управления и итогов
        self.create_summary_and_controls()

        # Блок управления Excel и AutoCAD
        self.create_excel_and_cad_controls()

    def create_order_info_section(self):
        """Создание секции с основной информацией."""
        grid = QGridLayout()

        # Поле для серийного номера (в левом верхнем углу)
        self.serial_number_input = QLineEdit()
        self.serial_number_input.setPlaceholderText("1111")  # Примерный начальный серийный номер
        self.serial_number_input.setFixedWidth(100)
        grid.addWidget(self.serial_number_input, 0, 0)

        # Поля ввода
        grid.addWidget(QLabel("Фирма:"), 1, 0)
        self.company_input = QComboBox()
        grid.addWidget(self.company_input, 1, 1)

        grid.addWidget(QLabel("Ответственное лицо:"), 1, 2)
        self.responsible_input = QLineEdit()
        grid.addWidget(self.responsible_input, 1, 3)

        grid.addWidget(QLabel("Телефон:"), 2, 0)
        self.phone_input = QLineEdit()
        grid.addWidget(self.phone_input, 2, 1)

        grid.addWidget(QLabel("Дата начала:"), 2, 2)
        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)
        self.start_date.setDate(QDate.currentDate())
        grid.addWidget(self.start_date, 2, 3)

        grid.addWidget(QLabel("Дата окончания:"), 3, 0)
        self.end_date = QDateEdit()
        self.end_date.setCalendarPopup(True)
        self.end_date.setDate(QDate.currentDate())
        grid.addWidget(self.end_date, 3, 1)

        grid.addWidget(QLabel("Адрес:"), 3, 2)
        self.address_input = QLineEdit()
        grid.addWidget(self.address_input, 3, 3)

        self.main_layout.addLayout(grid)
    
    def create_product_table(self):
        """Создание таблицы для ввода информации о продуктах."""
        # Создаем основной горизонтальный макет для таблицы и кнопок
        table_layout = QHBoxLayout()

        # Таблица продуктов
        self.table = QTableWidget(0, 6)  # Столбцы: Название, Ед. изм., Кол-во, Цена, Сумма, Примечание
        self.table.setHorizontalHeaderLabels([
            "Название продукта", "Ед. изм.", "Кол-во", "Цена", "Сумма", "Примечание"
        ])
        self.table.horizontalHeader().setStretchLastSection(True)
        table_layout.addWidget(self.table)

        # Устанавливаем делегат для числовых столбцов
        numeric_delegate = NumericDelegate(self)
        self.table.setItemDelegateForColumn(2, numeric_delegate)  # "Кол-во"
        self.table.setItemDelegateForColumn(3, numeric_delegate)  # "Цена"
        self.table.setItemDelegateForColumn(4, NumericDelegate(self))
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 4) or QTableWidgetItem("0.00")
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Запрет редактирования
            self.table.setItem(row, 4, item)
        # Боковая панель для кнопок
        button_layout = QVBoxLayout()

        # Кнопка "+" для добавления строки
        add_row_btn = QPushButton("+")
        add_row_btn.setToolTip("Добавить строку")
        add_row_btn.setFixedSize(30, 30)  # Размер кнопки
        add_row_btn.setStyleSheet("QPushButton { border-radius: 15px; background-color: #4CAF50; color: white; }")
        add_row_btn.clicked.connect(self.add_row)
        button_layout.addWidget(add_row_btn, alignment=Qt.AlignTop)

        # Кнопка "-" для удаления строки
        delete_row_btn = QPushButton("-")
        delete_row_btn.setToolTip("Удалить строку")
        delete_row_btn.setFixedSize(30, 30)  # Размер кнопки
        delete_row_btn.setStyleSheet("QPushButton { border-radius: 15px; background-color: #F44336; color: white; }")
        delete_row_btn.clicked.connect(self.delete_row)
        button_layout.addWidget(delete_row_btn, alignment=Qt.AlignTop)

        # Добавляем пространство для кнопок, чтобы они были сверху
        button_layout.addStretch()
        table_layout.addLayout(button_layout)

        # Добавляем макет в общий макет
        self.main_layout.addLayout(table_layout)


    def create_image_section(self):
        """Создание секции для загрузки и отображения изображений."""
        self.image_layout = QVBoxLayout()

        # Кнопка загрузки изображений
        upload_button = QPushButton("Загрузить изображение")
        upload_button.clicked.connect(self.upload_image)
        self.image_layout.addWidget(upload_button)

        # Горизонтальный макет для изображений
        self.image_display = QScrollArea()
        self.image_display.setWidgetResizable(True)
        self.image_display_widget = QWidget()
        self.image_display_layout = QHBoxLayout()  # Изображения будут располагаться в один ряд
        self.image_display_widget.setLayout(self.image_display_layout)
        self.image_display.setWidget(self.image_display_widget)

        self.image_layout.addWidget(self.image_display)
        self.main_layout.addLayout(self.image_layout)


    def create_summary_and_controls(self):
        """Создание блока итогов и кнопок управления."""
        summary_layout = QHBoxLayout()

        # Итоговая сумма
        self.total_label = QLabel("Итог: 0 AZN")
        summary_layout.addWidget(self.total_label)

        # Комиссия
        summary_layout.addWidget(QLabel("Комиссия (%):"))
        self.commission_input = QSpinBox()
        self.commission_input.setRange(0, 100)
        self.commission_input.setValue(15)
        self.commission_input.valueChanged.connect(self.update_total)
        summary_layout.addWidget(self.commission_input)

        # Кнопки
        self.send_button = QPushButton("Отправить заказ")
        self.send_button.clicked.connect(self.send_order)
        summary_layout.addWidget(self.send_button)

        self.clear_button = QPushButton("Очистить форму")
        self.clear_button.clicked.connect(self.clear_form)
        summary_layout.addLayout(summary_layout)

        self.main_layout.addLayout(summary_layout)

    def create_excel_and_cad_controls(self):
        """Создание секции для работы с Excel и AutoCAD."""
        layout = QHBoxLayout()

        # Кнопка загрузки данных из Excel
        load_excel_btn = QPushButton("Загрузить из Excel")
        load_excel_btn.clicked.connect(self.load_excel_data)
        layout.addWidget(load_excel_btn)

        # Кнопка сохранения в Excel
        save_excel_btn = QPushButton("Сохранить в Excel")
        save_excel_btn.clicked.connect(self.save_to_excel)
        layout.addWidget(save_excel_btn)

        # Кнопка открытия файла AutoCAD
        open_cad_btn = QPushButton("Открыть AutoCAD файл")
        open_cad_btn.clicked.connect(self.open_cad_file)
        layout.addWidget(open_cad_btn)

        self.main_layout.addLayout(layout)

    def load_excel_data(self):
        """Загрузка данных из Excel в форму."""
        file_path, _ = QFileDialog.getOpenFileName(self, "Выберите Excel файл", "", "Excel Files (*.xlsx)")
        if not file_path:
            return

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            self.company_input.setCurrentText(sheet["A1"].value or "")
            self.responsible_input.setText(sheet["A2"].value or "")
            self.phone_input.setText(sheet["A3"].value or "")

            self.table.setRowCount(0)
            for row in sheet.iter_rows(min_row=2, max_col=6, values_only=True):
                current_row = self.table.rowCount()
                self.table.insertRow(current_row)
                for col, value in enumerate(row):
                    self.table.setItem(current_row, col, QTableWidgetItem(str(value)))

        except Exception as e:
            print(f"Ошибка при загрузке Excel: {e}")

    from openpyxl.styles import Alignment, Font, Border, Side

    def save_to_excel(self):
        """Сохранение данных формы в Excel с форматированием."""
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить в Excel", "", "Excel Files (*.xlsx)")
        if not file_path:
            return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Установка названий колонок
            headers = [
                "Serial", "Фирма", "Ответственное лицо", "Телефон", "Дата начала",
                "Дата окончания", "Адрес", "Название продукта", "Ед. изм.", "Кол-во",
                "Цена", "Сумма", "Примечание"
            ]


            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                # Форматирование заголовков
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Запись основной информации
            serial = self.serial_number_input.text() or "1111"  # Значение по умолчанию
            sheet.cell(row=2, column=1, value=serial)
            sheet.cell(row=2, column=2, value=self.company_input.currentText())
            sheet.cell(row=2, column=3, value=self.responsible_input.text())
            sheet.cell(row=2, column=4, value=self.phone_input.text())
            sheet.cell(row=2, column=5, value=self.start_date.text())
            sheet.cell(row=2, column=6, value=self.end_date.text())
            sheet.cell(row=2, column=7, value=self.address_input.text())

            # Добавление формата для ячеек основной информации
            for row in range(2, 3):  # Строка основной информации
                for col in range(1, 8):  # Колонки от 1 до 7
                    cell = sheet.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin")
                    )

            # Проверка на дублирование серийного номера
            serial_duplicate_count = 1
            for row in range(self.table.rowCount()):
                current_serial = serial
                if row > 0:  # Добавляем суффикс для последующих строк
                    current_serial += f"-{serial_duplicate_count}"
                    serial_duplicate_count += 1

                base_row = row + 3  # Начиная с третьей строки
                sheet.cell(row=base_row, column=1, value=serial)  # Serial
                sheet.cell(row=base_row, column=2, value=self.company_input.currentText())  # Фирма
                sheet.cell(row=base_row, column=3, value=self.responsible_input.text())  # Ответственное лицо
                sheet.cell(row=base_row, column=4, value=self.phone_input.text())  # Телефон
                sheet.cell(row=base_row, column=5, value=self.start_date.text())  # Дата начала
                sheet.cell(row=base_row, column=6, value=self.end_date.text())  # Дата окончания
                sheet.cell(row=base_row, column=7, value=self.address_input.text())  # Адрес
                for col_idx in range(self.table.columnCount()):
                    item = self.table.item(row, col_idx)
                    value = item.text() if item else ""
                    sheet.cell(row=base_row, column=col_idx + 8, value=value)  # Продуктовые данные

                # Форматирование строк данных
                for col in range(1, 8):
                    cell = sheet.cell(row=base_row, column=col)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin")
                    )

            workbook.save(file_path)
            print("Данные успешно сохранены.")

        except Exception as e:
            print(f"Ошибка сохранения в Excel: {e}")



    def open_cad_file(self):
        """Открытие файла AutoCAD."""
        file_path, _ = QFileDialog.getOpenFileName(self, "Выберите файл AutoCAD", "", "DWG Files (*.dwg)")
        if not file_path:
            return

        try:
            subprocess.Popen(file_path, shell=True)
        except Exception as e:
            print(f"Ошибка при открытии AutoCAD файла: {e}")

    def add_row(self):
        """Добавить строку в таблицу."""
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)

    def delete_row(self):
        """Удалить выбранную строку из таблицы."""
        current_row = self.table.currentRow()
        if current_row >= 0:
            self.table.removeRow(current_row)


    def upload_image(self):
        """Загрузка изображения."""
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "Выберите изображение", "", "Images (*.png *.jpg *.jpeg)", options=options)
        if file_path:
            file_name = os.path.basename(file_path)
            save_path = os.path.join(IMAGES_FOLDER, file_name)
            QPixmap(file_path).save(save_path)

            # Создаем контейнер для изображения и кнопки
            container = QFrame()
            container.setStyleSheet("border: 1px solid #ccc; border-radius: 5px; margin: 5px;")
            container_layout = QVBoxLayout()
            container.setLayout(container_layout)

            # Добавляем изображение
            pixmap = QPixmap(save_path)
            label = QLabel()
            label.setPixmap(pixmap.scaled(150, 150, Qt.KeepAspectRatio))
            label.setAlignment(Qt.AlignCenter)
            container_layout.addWidget(label)

            # Добавляем кнопку удаления
            remove_button = QPushButton("Удалить")
            remove_button.setStyleSheet("""
                QPushButton {
                    background-color: #F44336;
                    color: white;
                    border: none;
                    padding: 5px;
                    border-radius: 5px;
                }
                QPushButton:hover {
                    background-color: #D32F2F;
                }
            """)
            remove_button.clicked.connect(lambda: self.remove_image(container, save_path))
            container_layout.addWidget(remove_button, alignment=Qt.AlignCenter)

            # Добавляем контейнер в горизонтальный лэйаут
            self.image_display_layout.addWidget(container)

    def remove_image(self, container, file_path):
        """Удаление изображения из виджета и файла."""
        # Удаляем виджет
        self.image_display_layout.removeWidget(container)
        container.deleteLater()

        # Удаляем файл изображения
        if os.path.exists(file_path):
            os.remove(file_path)

    def update_total(self):
        """Обновить итоговую сумму с учетом комиссии."""
        total = 0
        for row in range(self.table.rowCount()):
            try:
                # Получаем значения количества и цены
                quantity_item = self.table.item(row, 2)  # Кол-во
                price_item = self.table.item(row, 3)  # Цена
                
                # Проверяем, есть ли значения в ячейках
                quantity = float(quantity_item.text() if quantity_item and quantity_item.text() else 0)
                price = float(price_item.text() if price_item and price_item.text() else 0)
                
                # Рассчитываем сумму для строки
                row_sum = quantity * price
                total += row_sum
                
                # Обновляем ячейку "Сумма" (4-й столбец) и делаем её только для чтения
                item = QTableWidgetItem(f"{row_sum:.2f}")
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Запрет редактирования
                self.table.setItem(row, 4, item)
            except (ValueError, AttributeError):
                # Игнорируем строки с некорректными данными
                item = QTableWidgetItem("0.00")
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(row, 4, item)

        # Рассчитываем итоговую сумму с учетом комиссии
        commission = self.commission_input.value() / 100
        total_with_commission = total * (1 + commission)

        # Обновляем итоговый лейбл
        self.total_label.setText(f"Итог: {total_with_commission:.2f} AZN")



    def send_order(self):
        """Обработчик кнопки 'Отправить заказ'."""
        print("Заказ отправлен!")

    def clear_form(self):
        """Очистить форму."""
        self.company_input.clear()
        self.responsible_input.clear()
        self.phone_input.clear()
        self.address_input.clear()
        self.table.setRowCount(0)
        self.total_label.setText("Итог: 0 AZN")
        for i in reversed(range(self.image_display_layout.count())):
            widget = self.image_display_layout.itemAt(i).widget()
            widget.deleteLater()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = OrderApp()
    window.show()
    sys.exit(app.exec_())
